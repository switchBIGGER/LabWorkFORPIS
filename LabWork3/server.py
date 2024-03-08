import openpyxl
import pandas as pd


data = pd.read_csv('data.txt', names=['DateTime', 'Source', 'USD Rate'])

workbook = openpyxl.load_workbook('data.xlsx')

sheet = workbook.active

next_row = sheet.max_row + 1

for index, row in data.iterrows():
    sheet.cell(row=next_row, column=1).value = row['DateTime']
    sheet.cell(row=next_row, column=2).value = row['Source']
    sheet.cell(row=next_row, column=3).value = row['USD Rate']
    next_row += 1

workbook.save('data.xlsx')